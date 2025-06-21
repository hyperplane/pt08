import json
import base64
import io
from openpyxl import load_workbook
import logging
import uuid
import boto3
from datetime import datetime
from comment_analyzer import analyze_comments

logger = logging.getLogger()
logger.setLevel(logging.INFO)

s3_client = boto3.client('s3')
JOB_BUCKET = 'comment-analyzer-jobs'

def lambda_handler(event, context):
    headers = {
        'Content-Type': 'application/json'
    }
    
    try:
        if 'action' in event and event['action'] == 'process_async':
            job_id = event['job_id']
            logger.info(f"非同期処理開始: {job_id}")
            
            job_info = get_job_info(job_id)
            if not job_info:
                logger.error(f"ジョブ情報が見つかりません: {job_id}")
                return {'statusCode': 404, 'body': 'Job not found'}
            
            process_analysis_async(job_id, job_info)
            return {'statusCode': 200, 'body': 'OK'}
        
        if 'httpMethod' in event:
            method = event['httpMethod']
            body_str = event.get('body', '{}')
        elif 'requestContext' in event and 'http' in event['requestContext']:
            method = event['requestContext']['http']['method']
            body_str = event.get('body', '{}')
        else:
            method = 'POST'
            body_str = json.dumps(event)
        
        try:
            if isinstance(body_str, str):
                body = json.loads(body_str) if body_str else {}
            else:
                body = body_str
        except json.JSONDecodeError:
            return {
                'statusCode': 400,
                'headers': headers,
                'body': json.dumps({'error': 'Invalid JSON format'})
            }
        
        if body.get('start_job'):
            return start_analysis_job(body, headers)
        elif body.get('get_status'):
            return get_job_status(body, headers)
        elif body.get('get_result'):
            return get_job_result(body, headers)
        else:
            return process_sync_analysis(body, headers)
            
    except Exception as e:
        logger.error(f"Lambda handler error: {str(e)}")
        return {
            'statusCode': 500,
            'headers': headers,
            'body': json.dumps({'error': 'Internal server error'})
        }

def start_analysis_job(body, headers):
    try:
        if 'file_data' not in body:
            return {
                'statusCode': 400,
                'headers': headers,
                'body': json.dumps({'error': 'file_dataが見つかりません'})
            }
        
        test_mode = body.get('test_mode', False)
        job_id = str(uuid.uuid4())
        
        file_key = f"temp/{job_id}_file.xlsx"
        try:
            s3_client.put_object(
                Bucket=JOB_BUCKET,
                Key=file_key,
                Body=base64.b64decode(body['file_data']),
                ContentType='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
        except Exception as e:
            logger.error(f"ファイル一時保存エラー: {str(e)}")
            return {
                'statusCode': 500,
                'headers': headers,
                'body': json.dumps({'error': 'ファイルの一時保存に失敗しました'})
            }
        
        job_info = {
            'job_id': job_id,
            'status': 'started',
            'progress': 0,
            'total_comments': 0,
            'processed_comments': 0,
            'message': '分析を開始しています...',
            'created_at': datetime.now().isoformat(),
            'updated_at': datetime.now().isoformat(),
            'file_key': file_key,
            'test_mode': test_mode
        }
        
        save_job_info(job_id, job_info)
        logger.info(f"ジョブ開始: {job_id}")
        
        lambda_client = boto3.client('lambda', region_name='ap-northeast-1')
        
        lambda_client.invoke(
            FunctionName='comment-analyzer',
            InvocationType='Event',
            Payload=json.dumps({
                'action': 'process_async',
                'job_id': job_id
            })
        )
        
        return {
            'statusCode': 200,
            'headers': headers,
            'body': json.dumps({
                'job_id': job_id,
                'status': 'started',
                'message': '分析を開始しました。進捗はjob_idで確認できます。'
            })
        }
        
    except Exception as e:
        logger.error(f"ジョブ開始エラー: {str(e)}")
        return {
            'statusCode': 500,
            'headers': headers,
            'body': json.dumps({'error': 'ジョブ開始に失敗しました'})
        }

def get_job_status(body, headers):
    try:
        job_id = body.get('job_id')
        if not job_id:
            return {
                'statusCode': 400,
                'headers': headers,
                'body': json.dumps({'error': 'job_idが必要です'})
            }
        
        job_info = get_job_info(job_id)
        if not job_info:
            return {
                'statusCode': 404,
                'headers': headers,
                'body': json.dumps({'error': 'ジョブが見つかりません'})
            }
        
        return {
            'statusCode': 200,
            'headers': headers,
            'body': json.dumps({
                'job_id': job_id,
                'status': job_info['status'],
                'progress': job_info.get('progress', 0),
                'total_comments': job_info.get('total_comments', 0),
                'processed_comments': job_info.get('processed_comments', 0),
                'message': job_info.get('message', ''),
                'updated_at': job_info.get('updated_at', job_info.get('created_at'))
            })
        }
        
    except Exception as e:
        logger.error(f"ジョブ状況取得エラー: {str(e)}")
        return {
            'statusCode': 500,
            'headers': headers,
            'body': json.dumps({'error': 'ジョブ状況の取得に失敗しました'})
        }

def get_job_result(body, headers):
    try:
        job_id = body.get('job_id')
        if not job_id:
            return {
                'statusCode': 400,
                'headers': headers,
                'body': json.dumps({'error': 'job_idが必要です'})
            }
        
        job_info = get_job_info(job_id)
        if not job_info:
            return {
                'statusCode': 404,
                'headers': headers,
                'body': json.dumps({'error': 'ジョブが見つかりません'})
            }
        
        if job_info['status'] != 'completed':
            return {
                'statusCode': 400,
                'headers': headers,
                'body': json.dumps({'error': 'ジョブがまだ完了していません'})
            }
        
        return {
            'statusCode': 200,
            'headers': headers,
            'body': json.dumps(job_info.get('result', {}))
        }
        
    except Exception as e:
        logger.error(f"結果取得エラー: {str(e)}")
        return {
            'statusCode': 500,
            'headers': headers,
            'body': json.dumps({'error': '結果の取得に失敗しました'})
        }

def process_sync_analysis(body, headers):
    try:
        if 'file_data' not in body:
            return {
                'statusCode': 400,
                'headers': headers,
                'body': json.dumps({'error': 'file_dataが見つかりません'})
            }
        
        test_mode = body.get('test_mode', False)
        file_data = base64.b64decode(body['file_data'])
        result = analyze_comments(file_data, test_mode)
        
        return {
            'statusCode': 200,
            'headers': headers,
            'body': json.dumps(result)
        }
        
    except Exception as e:
        logger.error(f"同期分析エラー: {str(e)}")
        return {
            'statusCode': 500,
            'headers': headers,
            'body': json.dumps({'error': '分析に失敗しました'})
        }

def process_analysis_async(job_id, job_info):
    try:
        job_info['status'] = 'processing'
        job_info['updated_at'] = datetime.now().isoformat()
        save_job_info(job_id, job_info)
        
        file_key = job_info.get('file_key')
        if not file_key:
            raise ValueError("ファイルキーが見つかりません")
        
        try:
            response = s3_client.get_object(
                Bucket=JOB_BUCKET,
                Key=file_key
            )
            file_data = response['Body'].read()
        except Exception as e:
            raise ValueError(f"ファイルデータの取得に失敗しました: {str(e)}")
        
        def progress_callback(processed, total, message=""):
            job_info['processed_comments'] = processed
            job_info['total_comments'] = total
            job_info['progress'] = int((processed / total) * 100) if total > 0 else 0
            job_info['message'] = message
            job_info['updated_at'] = datetime.now().isoformat()
            save_job_info(job_id, job_info)
        
        result = analyze_comments(file_data, job_info['test_mode'], progress_callback)
        
        job_info['status'] = 'completed'
        job_info['result'] = result
        job_info['progress'] = 100
        job_info['message'] = '分析が完了しました'
        job_info['updated_at'] = datetime.now().isoformat()
        
        try:
            s3_client.delete_object(
                Bucket=JOB_BUCKET,
                Key=file_key
            )
            job_info.pop('file_key', None)
        except Exception as e:
            logger.warning(f"一時ファイル削除エラー: {str(e)}")
        
        save_job_info(job_id, job_info)
        
    except Exception as e:
        logger.error(f"非同期分析エラー: {str(e)}")
        job_info['status'] = 'error'
        job_info['error'] = str(e)
        job_info['message'] = '分析中にエラーが発生しました'
        job_info['updated_at'] = datetime.now().isoformat()
        
        file_key = job_info.get('file_key')
        if file_key:
            try:
                s3_client.delete_object(
                    Bucket=JOB_BUCKET,
                    Key=file_key
                )
                job_info.pop('file_key', None)
            except Exception as cleanup_e:
                logger.warning(f"エラー時一時ファイル削除エラー: {str(cleanup_e)}")
        
        save_job_info(job_id, job_info)

def save_job_info(job_id, job_info):
    try:
        if job_info['status'] in ['completed', 'error']:
            job_info_copy = job_info.copy()
            job_info_copy.pop('file_data', None)
            job_info_copy.pop('file_key', None)
            job_info = job_info_copy
        
        s3_client.put_object(
            Bucket=JOB_BUCKET,
            Key=f"jobs/{job_id}.json",
            Body=json.dumps(job_info, ensure_ascii=False),
            ContentType='application/json'
        )
    except Exception as e:
        logger.error(f"ジョブ情報保存エラー: {str(e)}")

def get_job_info(job_id):
    try:
        response = s3_client.get_object(
            Bucket=JOB_BUCKET,
            Key=f"jobs/{job_id}.json"
        )
        return json.loads(response['Body'].read().decode('utf-8'))
    except Exception as e:
        logger.error(f"ジョブ情報取得エラー: {str(e)}")
        return None

def load_excel_data(file_content):
    try:
        if len(file_content) < 100:
            raise ValueError("ファイルサイズが小さすぎます。有効なExcelファイルを選択してください。")
        
        wb = load_workbook(io.BytesIO(file_content))
        ws = wb.active
        
        if ws.max_row < 2:
            raise ValueError("Excelファイルにデータが見つかりません。最低2行（ヘッダー行+データ行）が必要です。")
        
        max_col = ws.max_column
        if max_col < 1:
            raise ValueError("Excelファイルに列が見つかりません。")
        
        comment_cols = max(1, max_col - 6) if max_col >= 7 else 1
        
        headers = {}
        for col in range(comment_cols, max_col + 1):
            header_cell = ws.cell(row=1, column=col)
            if header_cell.value:
                headers[col] = str(header_cell.value).strip()
            else:
                headers[col] = f'質問{col}'
        
        comments = []
        
        for row in ws.iter_rows(min_row=2, min_col=comment_cols, max_col=max_col):
            for cell in row:
                if cell.value and isinstance(cell.value, str) and len(cell.value.strip()) >= 10:
                    question_title = headers.get(cell.column, f'質問{cell.column}')
                    
                    comments.append({
                        'row_id': cell.row,
                        'column_name': question_title,
                        'comment': cell.value.strip()
                    })
        
        return comments
        
    except Exception as e:
        error_msg = str(e)
        if "not a zip file" in error_msg.lower():
            raise ValueError("無効なExcelファイル形式です。.xlsx または .xls ファイルを選択してください。")
        elif "bad zip file" in error_msg.lower():
            raise ValueError("Excelファイルが破損している可能性があります。別のファイルを試してください。")
        elif "no such file" in error_msg.lower():
            raise ValueError("ファイルが見つかりません。ファイルを再選択してください。")
        else:
            raise ValueError(f"Excelファイルの読み込みに失敗しました: {error_msg}")

def calculate_statistics(results):
    total = len(results)
    positive_count = sum(1 for r in results if r['sentiment'] == 'positive')
    negative_count = sum(1 for r in results if r['sentiment'] == 'negative')
    neutral_count = sum(1 for r in results if r['sentiment'] == 'neutral')
    danger_count = sum(1 for r in results if r['is_dangerous'])
    
    categories = {}
    for r in results:
        cat = r['category']
        categories[cat] = categories.get(cat, 0) + 1
    
    return {
        'total': total,
        'positive': positive_count,
        'negative': negative_count,
        'neutral': neutral_count,
        'dangerous': danger_count,
        'categories': categories
    } 