import logging
import boto3
import json
import re
import io
from openpyxl import load_workbook
from typing import List, Dict, Tuple

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

class CommentAnalyzer:
    def __init__(self):
        self.bedrock_client = boto3.client('bedrock-runtime', region_name='ap-northeast-1')
        self.model_id = "anthropic.claude-instant-v1"
        
    def analyze_comments_lambda(self, comments: list, progress_callback=None) -> list:
        results = []
        total_comments = len(comments)
        
        logger.info(f'AWS Bedrock分析開始: {total_comments}件')
        
        batch_size = 2
        processed_count = 0
        
        for i in range(0, total_comments, batch_size):
            batch = comments[i:i + batch_size]
            batch_results = self._analyze_batch_with_bedrock(batch, i)
            results.extend(batch_results)
            
            processed_count += len(batch)
            actual_processed = min(processed_count, total_comments)
            
            if progress_callback:
                progress_callback(actual_processed, total_comments, f"分析進捗: {actual_processed}/{total_comments}件完了")
            
            logger.info(f'分析進捗: {actual_processed}/{total_comments}件完了')
        
        logger.info(f'分析完了: {len(results)}件')
        return results
    
    def _analyze_batch_with_bedrock(self, batch_comments: list, start_index: int) -> list:
        try:
            prompt = self._build_analysis_prompt(batch_comments)
            
            response = self.bedrock_client.invoke_model(
                modelId=self.model_id,
                body=json.dumps({
                    "prompt": f"Human: {prompt}\n\nAssistant:",
                    "max_tokens_to_sample": 4000,
                    "temperature": 0.1,
                    "top_p": 0.9,
                    "stop_sequences": ["Human:", "Assistant:"]
                })
            )
            
            response_body = json.loads(response['body'].read())
            analysis_text = response_body.get('completion', '')
            
            if not analysis_text or analysis_text.strip() == "":
                logger.error("Bedrockからの応答が空です")
                return self._create_default_results(batch_comments, start_index)
            
            return self._parse_batch_results(analysis_text, batch_comments, start_index)
            
        except Exception as e:
            logger.error(f"Bedrock分析エラー: {e}")
            return self._create_default_results(batch_comments, start_index)
    
    def _build_analysis_prompt(self, batch_comments: list) -> str:
        comments_text = ""
        for i, comment_data in enumerate(batch_comments):
            comment = str(comment_data['comment']).strip()
            comments_text += f"{i+1}. {comment}\n"
        
        prompt = f"""以下の講義コメントを大学運営の観点から分析してください。

{comments_text}

各コメントについて、以下の形式のJSON配列で回答してください:
[
  {{
    "sentiment": "positive",
    "sentiment_score": 0.8,
    "category": "講義内容",
    "category_confidence": 0.9,
    "is_dangerous": false,
    "danger_score": 0.1,
    "importance_score": 0.7
  }}
]

分析基準:
1. sentiment: "positive"(肯定的), "negative"(否定的), "neutral"(中立)
2. category: 必ず「講義内容」「講義資料」「運営」「その他」のいずれかを選択

3. 危険コメント判定基準（講義運営の観点）:
   以下に該当する場合はis_dangerous: true、danger_score: 0.7以上に設定:

   【誹謗中傷・人格攻撃】
   - 講師・教員への人格否定、侮辱、悪口
   - 他の学生への攻撃的言及
   - 個人の外見・能力・人格への中傷

   【運営への攻撃・悪口】
   - 大学・学部・学科・授業運営への不当な攻撃
   - 運営方針への建設的でない批判
   - システムや制度への破壊的批判

   【運営妨害行為】
   - 授業妨害を示唆する内容
   - 他学生の学習を妨げる行為の示唆
   - 破壊的・反社会的行為の提案

   【不適切な言葉遣い】
   - 汚い言葉、罵詈雑言
   - 差別的表現、ヘイトスピーチ
   - 暴力的・脅迫的表現
   - 性的・わいせつな表現
   - 「死ね」「殺す」「バカ」「クソ」「ムカつく」「最悪」「ひどい」「うざい」など

4. danger_score設定:
   - 0.9-1.0: 緊急対応必要（脅迫・暴力的表現）
   - 0.8-0.9: 即座に対応必要（人格攻撃・誹謗中傷）
   - 0.7-0.8: 要注意（運営攻撃・不適切言葉）
   - 0.5-0.7: 注意（建設的でない強い批判）
   - 0.0-0.5: 安全（建設的意見・要望）

5. importance_score: コメントの重要度（0.0-1.0）

注意：「つまらない」「わからない」「嫌い」「やめたい」程度でも、文脈や表現によっては危険度0.5-0.7として判定してください。攻撃的な表現や感情的な否定は積極的に危険と判定してください。"""
        
        return prompt
    
    def _parse_batch_results(self, analysis_text: str, batch_comments: list, start_index: int) -> list:
        results = []
        
        try:
            json_patterns = [
                r'\[[\s\S]*?\]',
                r'\{[\s\S]*?\}',
                r'```json\s*([\s\S]*?)\s*```',
                r'```\s*([\s\S]*?)\s*```'
            ]
            
            json_str = None
            for pattern in json_patterns:
                match = re.search(pattern, analysis_text, re.DOTALL)
                if match:
                    if 'json' in pattern:
                        json_str = match.group(1).strip()
                    else:
                        json_str = match.group(0).strip()
                    break
            
            if not json_str:
                logger.warning("JSON形式が見つかりません")
                return self._create_default_results(batch_comments, start_index)
            
            if json_str.startswith('{') and json_str.endswith('}'):
                json_str = f"[{json_str}]"
            
            parsed_results = json.loads(json_str)
            
            if len(parsed_results) != len(batch_comments):
                while len(parsed_results) < len(batch_comments):
                    parsed_results.append({
                        'sentiment': 'neutral',
                        'sentiment_score': 0.5,
                        'category': 'その他',
                        'category_confidence': 0.5,
                        'is_dangerous': False,
                        'danger_score': 0.1,
                        'importance_score': 0.5
                    })
            
            for i, (result, comment_data) in enumerate(zip(parsed_results, batch_comments)):
                structured_result = {
                    'comment': str(comment_data['comment']).strip(),
                    'row_id': comment_data.get('row_id', start_index + i),
                    'column_name': comment_data.get('column_name', 'comment'),
                    'sentiment': result.get('sentiment', 'neutral'),
                    'sentiment_score': float(result.get('sentiment_score', 0.5)),
                    'category': result.get('category', 'その他'),
                    'category_confidence': float(result.get('category_confidence', 0.5)),
                    'is_dangerous': bool(result.get('is_dangerous', False)),
                    'danger_score': float(result.get('danger_score', 0.1)),
                    'danger_reasons': f"危険度: {result.get('danger_score', 0.1)}",
                    'importance_score': float(result.get('importance_score', 0.5)),
                    'specificity_score': 0.8,
                    'urgency_score': 0.7,
                    'commonality_score': 0.6
                }
                results.append(structured_result)
                
        except json.JSONDecodeError as e:
            logger.error(f"JSON解析エラー: {e}")
            results = self._create_default_results(batch_comments, start_index)
        except Exception as e:
            logger.error(f"結果パースエラー: {e}")
            results = self._create_default_results(batch_comments, start_index)
            
        return results
    
    def _create_default_results(self, batch_comments: list, start_index: int) -> list:
        results = []
        for i, comment_data in enumerate(batch_comments):
            result = {
                'comment': str(comment_data['comment']).strip(),
                'row_id': comment_data.get('row_id', start_index + i),
                'column_name': comment_data.get('column_name', 'comment'),
                'sentiment': 'neutral',
                'sentiment_score': 0.5,
                'category': 'その他',
                'category_confidence': 0.5,
                'is_dangerous': False,
                'danger_score': 0.1,
                'danger_reasons': "分析エラー",
                'importance_score': 0.3,
                'specificity_score': 0.5,
                'urgency_score': 0.5,
                'commonality_score': 0.5
            }
            results.append(result)
        return results

def analyze_comments(file_data, test_mode=False, progress_callback=None):
    try:
        comments = load_excel_data(file_data)
        logger.info(f"読み込み完了: {len(comments)}件のコメント")
        
        if not comments:
            raise ValueError("コメントが見つかりませんでした。Excelファイルの内容を確認してください。")
        
        if test_mode:
            comments = comments[:100]
            logger.info("テストモード: 100件のみ処理")
        
        total_comments = len(comments)
        
        if progress_callback:
            progress_callback(0, total_comments, "分析を開始しています...")
        
        analyzer = CommentAnalyzer()
        logger.info(f"分析開始: {total_comments}件のコメントを処理します")
        
        results = analyzer.analyze_comments_lambda(comments, progress_callback)
        
        logger.info(f"分析処理完了: {len(results)}件の結果を生成")
        
        if progress_callback:
            progress_callback(total_comments, total_comments, "統計情報を計算中...")
        
        stats = calculate_statistics(results)
        logger.info(f"統計計算完了: total={stats['total']}, positive={stats['positive']}, negative={stats['negative']}")
        
        response_data = {
            'success': True,
            'message': f'分析が完了しました（{len(results)}件処理）',
            'statistics': stats,
            'results': results
        }
        
        return response_data
                
    except Exception as e:
        logger.error(f"分析エラー: {str(e)}")
        raise e

def load_excel_data(file_content):
    try:
        if len(file_content) < 100:
            raise ValueError("ファイルサイズが小さすぎます。有効なExcelファイルを選択してください。")
        
        wb = load_workbook(io.BytesIO(file_content))
        ws = wb.active
        
        if ws.max_row < 2:
            raise ValueError("Excelファイルにデータが見つかりません。最低2行（ヘッダー行+データ行）が必要です。")
        
        max_col = ws.max_column
        if max_col < 1:
            raise ValueError("Excelファイルに列が見つかりません。")
        
        comment_cols = max(1, max_col - 6) if max_col >= 7 else 1
        
        headers = {}
        for col in range(comment_cols, max_col + 1):
            header_cell = ws.cell(row=1, column=col)
            if header_cell.value:
                headers[col] = str(header_cell.value).strip()
            else:
                headers[col] = f'質問{col}'
        
        comments = []
        
        for row in ws.iter_rows(min_row=2, min_col=comment_cols, max_col=max_col):
            for cell in row:
                if cell.value and isinstance(cell.value, str) and len(cell.value.strip()) >= 10:
                    question_title = headers.get(cell.column, f'質問{cell.column}')
                    
                    comments.append({
                        'row_id': cell.row,
                        'column_name': question_title,
                        'comment': cell.value.strip()
                    })
        
        return comments
        
    except Exception as e:
        error_msg = str(e)
        if "not a zip file" in error_msg.lower():
            raise ValueError("無効なExcelファイル形式です。.xlsx または .xls ファイルを選択してください。")
        elif "bad zip file" in error_msg.lower():
            raise ValueError("Excelファイルが破損している可能性があります。別のファイルを試してください。")
        elif "no such file" in error_msg.lower():
            raise ValueError("ファイルが見つかりません。ファイルを再選択してください。")
        else:
            raise ValueError(f"Excelファイルの読み込みに失敗しました: {error_msg}")

def calculate_statistics(results):
    total = len(results)
    positive_count = sum(1 for r in results if r['sentiment'] == 'positive')
    negative_count = sum(1 for r in results if r['sentiment'] == 'negative')
    neutral_count = sum(1 for r in results if r['sentiment'] == 'neutral')
    danger_count = sum(1 for r in results if r['is_dangerous'])
    
    categories = {}
    for r in results:
        cat = r['category']
        categories[cat] = categories.get(cat, 0) + 1
    
    return {
        'total': total,
        'positive': positive_count,
        'negative': negative_count,
        'neutral': neutral_count,
        'dangerous': danger_count,
        'categories': categories
    }
